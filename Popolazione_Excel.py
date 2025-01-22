import openpyxl # libreria utilizzata per poter lavorare con i fogli di calcolo Excel
from openpyxl.styles import Font, Alignment # importiamo le classi Font e Alignment dalla libreria openpyxl
import faker # libreria utilizzata per generare dati fittizzi
import random # libreria utilizzata per generare numeri casuali
import os # libreria utilizzata per interagire con il sistema operativo

def creazione_file_excel(nome_file):
    try:
        # Verifica se il file esiste già
        if os.path.exists(nome_file):
            print(f"Il file {nome_file} esiste già.")
            return  # Esce dalla funzione se il file esiste

        # Creazione del file excel 
        file_excel = openpyxl.Workbook()

        # Accesso al foglio di lavoro
        foglio_excel = file_excel.active
        foglio_excel.title = "Dati_Utenti" # Titolo del foglio di lavoro

        # Titolo delle colonne
        titolo_colonne = ["Nome", "Cognome", "Email", "Numero di telefono"]
        foglio_excel.append(titolo_colonne)

       # Creazione dell'istanza Faker in localizzazione italiana per generare dati casuale per gli utenti
        dati_casuali = faker.Faker('it_IT')

        # Lista di domini email personalizzati
        domini_email = ['@libero.it', '@gmail.com', '@yahoo.it', '@hotmail.it', '@outlook.com']

        # Funzione per generare una email con dominio casuale
        def email_personalizzata(nome, cognome):
            dominio = random.choice(domini_email)  # Scelta casuale del dominio
            # Crea una email con nome e cognome (es. giuseppemorello@gmail.com)
            return f"{nome.lower()}{cognome.lower()}{dominio}"

        # Funzione per generare un numero di telefono italiano 
        def telefono_italiano():
            return f"+39 {random.randint(3000000000, 3999999999)}" # Generazione di un numero casuale tra i due range

        # Generazione dati utenti casuali e inserimento all'interno del foglio di lavoro
        dati_utenti = []
        for utente in range(10):  # Genera 10 utenti
            nome = dati_casuali.first_name()
            cognome = dati_casuali.last_name()
            email = email_personalizzata(nome, cognome)  # Usa la funzione per generare email personalizzate
            telefono = telefono_italiano()  # Usa la funzione per generare un numero di telefono italiano
            
            # Inserisce i dati generati all'interno della lista dati_utenti
            dati_utenti.append([nome, cognome, email, telefono])

            # Ordinamento della lista degli utenti in base ai nomi dalla a alla z
            dati_utenti.sort(key=lambda x: x[0])

        # Inserisce i dati all'interno del foglio
        for riga in dati_utenti:
            foglio_excel.append(riga)

        # Formattazione delle titolo_colonne
        # utilizzo del ciclo for per iterare ogni colonna contenente un'intestazione, ovvero sulle 4 colonne (A,B,C,D)
        for colonna in range(1, len(titolo_colonne) + 1): 
            cella = foglio_excel.cell(row=1, column=colonna) # accesso ad una cella specifica del foglio di calcolo
            cella.font = Font(bold=True)  # Testo in grassetto
            cella.alignment = Alignment(horizontal="center")  # Allineamento centrale del testo

        # Formattazione delle colonne sul foglio excel per evitare troncamenti
        for colonna in range(1, len(titolo_colonne) + 1):
            lunghezza_massima = 0 
            lettera_colonna = openpyxl.utils.get_column_letter(colonna) 
            for riga in foglio_excel.iter_rows(min_col=colonna, max_col=colonna):
                for cella in riga:
                    if cella.value:
                        lunghezza_massima = max(lunghezza_massima, len(str(cella.value)))
            # aggiunta di spazio extra per fare in modo che i valori non stiano troppo attacati al margine
            regolazione_larghezza = (lunghezza_massima + 2)  
            foglio_excel.column_dimensions[lettera_colonna].width = regolazione_larghezza 

        # Salvataggio file excel
        file_excel.save(nome_file)
        print(f"Il file Excel {nome_file} è stato creato con successo.")

    except Exception as e:
        print(f"Si è verificato un errore: {e}")

# Esecuzione della funzione
nome_file_excel = "Dati_Utenti.xlsx"
creazione_file_excel(nome_file_excel)